"""
Create Excel Templates for Model Registry
Generates Excel templates with proper formatting and validation rules
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Template configurations based on database schema
TEMPLATE_CONFIGS = {
    'model_type': {
        'columns': [
            {'name': 'TYPE_CODE', 'type': 'text', 'required': True, 'max_length': 20, 'description': 'Mã loại mô hình (PD, LGD, EAD, etc.)'},
            {'name': 'TYPE_NAME', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Tên đầy đủ của loại mô hình'},
            {'name': 'TYPE_DESCRIPTION', 'type': 'text', 'required': False, 'max_length': 500, 'description': 'Mô tả chi tiết về loại mô hình'},
            {'name': 'IS_ACTIVE', 'type': 'boolean', 'required': False, 'default': True, 'description': 'Trạng thái hoạt động (1=Active, 0=Inactive)'}
        ],
        'sample_data': [
            ['PD', 'Probability of Default', 'Mô hình ước tính xác suất vỡ nợ', 1],
            ['LGD', 'Loss Given Default', 'Mô hình ước tính tỷ lệ tổn thất', 1],
            ['EAD', 'Exposure at Default', 'Mô hình ước tính giá trị rủi ro', 1],
            ['B-SCORE', 'Behavioral Scorecard', 'Thẻ điểm đánh giá hành vi', 1],
            ['A-SCORE', 'Application Scorecard', 'Thẻ điểm đánh giá đăng ký', 1],
            ['C-SCORE', 'Collection Scorecard', 'Thẻ điểm đánh giá thu hồi nợ', 1],
            ['SEGMENT', 'Segmentation Model', 'Mô hình phân khúc khách hàng', 1],
            ['EARLY_WARN', 'Early Warning Signal', 'Mô hình cảnh báo sớm', 1],
            ['LIMIT', 'Limit Setting Model', 'Mô hình thiết lập hạn mức', 1],
            ['STRESS', 'Stress Testing Model', 'Mô hình kiểm tra căng thẳng', 1],
            ['PRICING', 'Risk-based Pricing', 'Mô hình định giá dựa trên rủi ro', 1],
            ['FRAUD', 'Fraud Detection', 'Mô hình phát hiện gian lận', 1]
        ]
    },
    
    'model_registry': {
        'columns': [
            {'name': 'TYPE_ID', 'type': 'number', 'required': True, 'description': 'ID loại mô hình (tham chiếu MODEL_TYPE)'},
            {'name': 'MODEL_NAME', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Tên mô hình'},
            {'name': 'MODEL_DESCRIPTION', 'type': 'text', 'required': False, 'max_length': 500, 'description': 'Mô tả chi tiết về mô hình'},
            {'name': 'MODEL_VERSION', 'type': 'text', 'required': True, 'max_length': 20, 'description': 'Phiên bản mô hình (1.0, 2.1, etc.)'},
            {'name': 'SOURCE_DATABASE', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Tên database chứa dữ liệu nguồn'},
            {'name': 'SOURCE_SCHEMA', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Schema chứa dữ liệu nguồn'},
            {'name': 'SOURCE_TABLE_NAME', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Tên bảng chứa dữ liệu nguồn'},
            {'name': 'REF_SOURCE', 'type': 'text', 'required': False, 'max_length': 255, 'description': 'Tài liệu tham khảo'},
            {'name': 'EFF_DATE', 'type': 'date', 'required': True, 'description': 'Ngày có hiệu lực (YYYY-MM-DD)'},
            {'name': 'EXP_DATE', 'type': 'date', 'required': True, 'description': 'Ngày hết hiệu lực (YYYY-MM-DD)'},
            {'name': 'IS_ACTIVE', 'type': 'boolean', 'required': False, 'default': True, 'description': 'Trạng thái hoạt động'},
            {'name': 'PRIORITY', 'type': 'number', 'required': False, 'default': 1, 'description': 'Độ ưu tiên (1=cao nhất)'},
            {'name': 'MODEL_CATEGORY', 'type': 'dropdown', 'required': False, 'options': ['Retail', 'SME', 'Corporate', 'Wholesale', 'Treasury', 'Market Risk', 'Operational Risk'], 'description': 'Phân loại mô hình'},
            {'name': 'SEGMENT_CRITERIA', 'type': 'text', 'required': False, 'description': 'Tiêu chí phân khúc (JSON format)'}
        ],
        'sample_data': [
            [1, 'PD_RETAIL', 'Mô hình xác suất vỡ nợ cho khách hàng cá nhân', '1.0', 'RISK_MODELS', 'dbo', 'PD_RETAIL_RESULTS', 'Risk Management/PD_RETAIL_v1.0.pdf', '2024-01-01', '2026-01-01', 1, 1, 'Retail', '{"customer_segment": "RETAIL"}'],
            [1, 'PD_SME', 'Mô hình xác suất vỡ nợ cho khách hàng SME', '2.0', 'RISK_MODELS', 'dbo', 'PD_SME_RESULTS', 'Risk Management/PD_SME_v2.0.pdf', '2024-04-01', '2026-04-01', 1, 1, 'SME', '{"customer_segment": "SME"}'],
            [2, 'LGD_RETAIL', 'Mô hình tổn thất khi vỡ nợ cho khách hàng cá nhân', '1.0', 'RISK_MODELS', 'dbo', 'LGD_RETAIL_RESULTS', 'Risk Management/LGD_RETAIL_v1.0.pdf', '2024-01-01', '2026-01-01', 1, 1, 'Retail', '{"customer_segment": "RETAIL"}'],
            [3, 'EAD_RETAIL', 'Mô hình giá trị rủi ro cho khách hàng cá nhân', '1.0', 'RISK_MODELS', 'dbo', 'EAD_RETAIL_RESULTS', 'Risk Management/EAD_RETAIL_v1.0.pdf', '2024-01-01', '2026-01-01', 1, 1, 'Retail', '{"customer_segment": "RETAIL"}'],
            [4, 'BSCORE_RETAIL', 'Thẻ điểm hành vi cho khách hàng cá nhân', '2.0', 'RISK_MODELS', 'dbo', 'BSCORE_RETAIL_RESULTS', 'Risk Management/BSCORE_RETAIL_v2.0.pdf', '2024-06-01', '2026-06-01', 1, 1, 'Retail', '{"customer_segment": "RETAIL"}']
        ]
    },
    
    'feature_registry': {
        'columns': [
            {'name': 'FEATURE_NAME', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Tên đầy đủ của đặc trưng'},
            {'name': 'FEATURE_CODE', 'type': 'text', 'required': True, 'max_length': 50, 'description': 'Mã ngắn gọn của đặc trưng'},
            {'name': 'FEATURE_DESCRIPTION', 'type': 'text', 'required': False, 'max_length': 500, 'description': 'Mô tả chi tiết về đặc trưng'},
            {'name': 'DATA_TYPE', 'type': 'dropdown', 'required': True, 'options': [
                'NUMERIC', 'CATEGORICAL', 'DATE', 'TEXT', 'BINARY', 'DECIMAL', 'INTEGER', 'FLOAT', 
                'DOUBLE', 'MONEY', 'CURRENCY', 'PERCENTAGE', 'RATIO', 'COUNT', 'DURATION', 'TIMESTAMP'
            ], 'description': 'Kiểu dữ liệu kỹ thuật trong hệ thống'},
            {'name': 'VALUE_TYPE', 'type': 'dropdown', 'required': True, 'options': [
                'CONTINUOUS', 'DISCRETE', 'BINARY', 'NOMINAL', 'ORDINAL', 'INTERVAL', 'RATIO', 
                'CYCLICAL', 'TEMPORAL', 'SPATIAL', 'HIERARCHICAL', 'COMPOSITE', 'DERIVED', 'AGGREGATED'
            ], 'description': 'Loại giá trị theo ý nghĩa nghiệp vụ'},
            {'name': 'SOURCE_SYSTEM', 'type': 'text', 'required': True, 'max_length': 100, 'description': 'Hệ thống nguồn của đặc trưng'},
            {'name': 'BUSINESS_CATEGORY', 'type': 'dropdown', 'required': False, 'options': [
                'DEMOGRAPHIC', 'FINANCIAL', 'BEHAVIORAL', 'RELATIONSHIP', 'ACCOUNT', 'BALANCE', 
                'DELINQUENCY', 'BUREAU', 'COLLATERAL', 'TRANSACTION', 'PRODUCT', 'GEOGRAPHIC', 
                'TEMPORAL', 'EXTERNAL', 'MARKET', 'REGULATORY', 'OPERATIONAL', 'TECHNICAL'
            ], 'description': 'Phân loại nghiệp vụ của đặc trưng'},
            {'name': 'DOMAIN_KNOWLEDGE', 'type': 'text', 'required': False, 'description': 'Kiến thức nghiệp vụ về đặc trưng'},
            {'name': 'IS_PII', 'type': 'boolean', 'required': False, 'default': False, 'description': 'Cờ thông tin nhận dạng cá nhân'},
            {'name': 'IS_SENSITIVE', 'type': 'boolean', 'required': False, 'default': False, 'description': 'Cờ dữ liệu nhạy cảm'},
            {'name': 'DEFAULT_VALUE', 'type': 'text', 'required': False, 'max_length': 100, 'description': 'Giá trị mặc định nếu thiếu'},
            {'name': 'VALID_MIN_VALUE', 'type': 'text', 'required': False, 'max_length': 100, 'description': 'Giá trị tối thiểu hợp lệ'},
            {'name': 'VALID_MAX_VALUE', 'type': 'text', 'required': False, 'max_length': 100, 'description': 'Giá trị tối đa hợp lệ'},
            {'name': 'VALID_VALUES', 'type': 'text', 'required': False, 'description': 'Mảng JSON chứa các giá trị hợp lệ'},
            {'name': 'BUSINESS_OWNER', 'type': 'text', 'required': False, 'max_length': 100, 'description': 'Phòng ban/người chịu trách nhiệm'},
            {'name': 'UPDATE_FREQUENCY', 'type': 'dropdown', 'required': False, 'options': [
                'REAL_TIME', 'HOURLY', 'DAILY', 'WEEKLY', 'MONTHLY', 'QUARTERLY', 'YEARLY', 'STATIC', 
                'ON_DEMAND', 'BATCH', 'STREAMING', 'EVENT_DRIVEN'
            ], 'description': 'Tần suất cập nhật dữ liệu'}
        ],
        'sample_data': [
            # Demographic Features
            ['Customer Age', 'CUST_AGE', 'Tuổi của khách hàng', 'NUMERIC', 'CONTINUOUS', 'CORE_BANKING', 'DEMOGRAPHIC', 'Tuổi là yếu tố quan trọng trong đánh giá rủi ro', False, False, None, '18', '100', None, 'Customer Data Team', 'MONTHLY'],
            ['Gender', 'GENDER', 'Giới tính của khách hàng', 'CATEGORICAL', 'NOMINAL', 'CORE_BANKING', 'DEMOGRAPHIC', 'Giới tính có thể ảnh hưởng đến hành vi tài chính', True, False, 'UNKNOWN', None, None, '["MALE", "FEMALE", "OTHER"]', 'Customer Data Team', 'STATIC'],
            ['Marital Status', 'MARITAL_STATUS', 'Tình trạng hôn nhân', 'CATEGORICAL', 'NOMINAL', 'CORE_BANKING', 'DEMOGRAPHIC', 'Tình trạng hôn nhân ảnh hưởng đến độ ổn định tài chính', False, False, 'UNKNOWN', None, None, '["SINGLE", "MARRIED", "DIVORCED", "WIDOWED"]', 'Customer Data Team', 'QUARTERLY'],
            ['Education Level', 'EDU_LEVEL', 'Trình độ học vấn', 'CATEGORICAL', 'ORDINAL', 'CORE_BANKING', 'DEMOGRAPHIC', 'Trình độ học vấn cao thường có thu nhập ổn định hơn', False, False, 'UNKNOWN', None, None, '["PRIMARY", "SECONDARY", "HIGH_SCHOOL", "COLLEGE", "UNIVERSITY", "POST_GRADUATE"]', 'Customer Data Team', 'STATIC'],
            
            # Financial Features
            ['Customer Income', 'INCOME', 'Thu nhập hàng tháng của khách hàng', 'NUMERIC', 'CONTINUOUS', 'CORE_BANKING', 'FINANCIAL', 'Thu nhập là yếu tố chính trong đánh giá khả năng trả nợ', False, True, '0', '0', '1000000000', None, 'Risk Analytics Team', 'QUARTERLY'],
            ['Income Category', 'INCOME_CAT', 'Phân loại thu nhập', 'CATEGORICAL', 'ORDINAL', 'DATA_WAREHOUSE', 'FINANCIAL', 'Phân loại thu nhập giúp phân đoạn khách hàng', False, True, 'MEDIUM', None, None, '["LOW", "MEDIUM", "HIGH", "VERY_HIGH"]', 'Risk Analytics Team', 'QUARTERLY'],
            ['Total Assets', 'TOTAL_ASSETS', 'Tổng tài sản của khách hàng', 'NUMERIC', 'CONTINUOUS', 'CORE_BANKING', 'FINANCIAL', 'Tổng tài sản thể hiện khả năng tài chính', False, True, '0', '0', '10000000000', None, 'Risk Analytics Team', 'QUARTERLY'],
            ['Debt to Income Ratio', 'DTI_RATIO', 'Tỷ lệ nợ trên thu nhập', 'NUMERIC', 'CONTINUOUS', 'DATA_WAREHOUSE', 'FINANCIAL', 'Tỷ lệ nợ trên thu nhập là chỉ báo quan trọng về rủi ro', False, True, '0', '0', '10', None, 'Risk Analytics Team', 'MONTHLY'],
            
            # Behavioral Features
            ['Payment History', 'PAYMENT_HIST', 'Lịch sử thanh toán', 'CATEGORICAL', 'ORDINAL', 'COLLECTION_SYSTEM', 'BEHAVIORAL', 'Lịch sử thanh toán là chỉ báo mạnh về hành vi tín dụng', False, False, 'UNKNOWN', None, None, '["EXCELLENT", "GOOD", "FAIR", "POOR", "VERY_POOR"]', 'Collections Team', 'DAILY'],
            ['Utilization Ratio', 'UTIL_RATIO', 'Tỷ lệ sử dụng hạn mức tín dụng', 'NUMERIC', 'CONTINUOUS', 'DATA_WAREHOUSE', 'BEHAVIORAL', 'Tỷ lệ sử dụng hạn mức cao thường liên quan đến rủi ro cao', False, False, '0', '0', '1', None, 'Risk Analytics Team', 'DAILY'],
            ['Transaction Frequency', 'TXN_FREQ', 'Tần suất giao dịch hàng tháng', 'NUMERIC', 'DISCRETE', 'CORE_BANKING', 'BEHAVIORAL', 'Tần suất giao dịch thể hiện mức độ hoạt động của tài khoản', False, False, '0', '0', '1000', None, 'Transaction Analytics Team', 'MONTHLY'],
            ['Average Transaction Amount', 'AVG_TXN_AMT', 'Số tiền giao dịch trung bình', 'NUMERIC', 'CONTINUOUS', 'CORE_BANKING', 'BEHAVIORAL', 'Số tiền giao dịch trung bình phản ánh quy mô hoạt động', False, False, '0', '0', '1000000000', None, 'Transaction Analytics Team', 'MONTHLY'],
            
            # Account Features
            ['Account Age', 'ACC_AGE', 'Tuổi tài khoản tính theo tháng', 'NUMERIC', 'DISCRETE', 'CORE_BANKING', 'ACCOUNT', 'Tuổi tài khoản chỉ ra thời gian hoạt động của sản phẩm', False, False, '0', '0', '600', None, 'Account Management Team', 'MONTHLY'],
            ['Product Type', 'PROD_TYPE', 'Loại sản phẩm tài chính', 'CATEGORICAL', 'NOMINAL', 'CORE_BANKING', 'PRODUCT', 'Loại sản phẩm có các đặc điểm rủi ro khác nhau', False, False, None, None, None, '["MORTGAGE", "CREDIT_CARD", "PERSONAL_LOAN", "AUTO_LOAN", "SECURED_LOAN", "UNSECURED_LOAN"]', 'Product Management Team', 'STATIC'],
            ['Credit Limit', 'CREDIT_LIMIT', 'Hạn mức tín dụng được cấp', 'NUMERIC', 'CONTINUOUS', 'CORE_BANKING', 'ACCOUNT', 'Hạn mức tín dụng thể hiện mức độ rủi ro mà ngân hàng sẵn sàng chấp nhận', False, False, '0', '0', '1000000000', None, 'Credit Risk Management Team', 'MONTHLY'],
            ['Interest Rate', 'INT_RATE', 'Lãi suất áp dụng cho tài khoản', 'NUMERIC', 'CONTINUOUS', 'CORE_BANKING', 'ACCOUNT', 'Lãi suất thường phản ánh mức độ rủi ro được đánh giá', False, False, '0', '0', '50', None, 'Product Management Team', 'MONTHLY'],
            
            # Delinquency Features
            ['Days Past Due', 'DPD', 'Số ngày quá hạn hiện tại', 'NUMERIC', 'DISCRETE', 'COLLECTION_SYSTEM', 'DELINQUENCY', 'DPD là thước đo trực tiếp về hành vi thanh toán', False, False, '0', '0', '180', None, 'Collections Team', 'DAILY'],
            ['DPD 30+ Last 12M', 'DPD_30_L12M', 'Số lần quá hạn trên 30 ngày trong 12 tháng qua', 'NUMERIC', 'DISCRETE', 'DATA_WAREHOUSE', 'DELINQUENCY', 'Tần suất quá hạn là chỉ báo mạnh mẽ về hành vi thanh toán', False, False, '0', '0', '12', None, 'Risk Analytics Team', 'MONTHLY'],
            ['Maximum DPD Last 12M', 'MAX_DPD_L12M', 'Số ngày quá hạn tối đa trong 12 tháng qua', 'NUMERIC', 'DISCRETE', 'DATA_WAREHOUSE', 'DELINQUENCY', 'Mức độ quá hạn tối đa trong quá khứ là chỉ báo về mức độ nghiêm trọng', False, False, '0', '0', '180', None, 'Risk Analytics Team', 'MONTHLY'],
            ['Default History', 'DEFAULT_HIST', 'Lịch sử vỡ nợ', 'NUMERIC', 'DISCRETE', 'COLLECTION_SYSTEM', 'DELINQUENCY', 'Lịch sử vỡ nợ là một trong những chỉ báo mạnh nhất về rủi ro', False, False, '0', '0', '10', None, 'Collections Team', 'MONTHLY'],
            
            # Bureau Features
            ['Bureau Score', 'BUREAU_SCORE', 'Điểm tín dụng từ cục thông tin tín dụng', 'NUMERIC', 'DISCRETE', 'CREDIT_BUREAU', 'BUREAU', 'Điểm tín dụng bên ngoài cung cấp đánh giá khách quan', False, False, None, '300', '900', None, 'Credit Bureau Team', 'MONTHLY'],
            ['Credit Utilization External', 'CREDIT_UTIL_EXT', 'Tỷ lệ sử dụng tín dụng tổng thể', 'NUMERIC', 'CONTINUOUS', 'CREDIT_BUREAU', 'BUREAU', 'Tỷ lệ sử dụng tín dụng bên ngoài cho thấy mức độ sử dụng tín dụng', False, False, None, '0', '1', None, 'Credit Bureau Team', 'MONTHLY'],
            ['Inquiries Last 6M', 'INQ_L6M', 'Số lần truy vấn thông tin tín dụng trong 6 tháng qua', 'NUMERIC', 'DISCRETE', 'CREDIT_BUREAU', 'BUREAU', 'Số lượng truy vấn gần đây có thể chỉ ra khách hàng đang tìm kiếm tín dụng', False, False, '0', '0', '50', None, 'Credit Bureau Team', 'MONTHLY'],
            ['Credit History Length', 'CREDIT_HIST_LEN', 'Độ dài lịch sử tín dụng tính theo tháng', 'NUMERIC', 'DISCRETE', 'CREDIT_BUREAU', 'BUREAU', 'Lịch sử tín dụng dài hơn thường chỉ ra hồ sơ tín dụng thành thục hơn', False, False, '0', '0', '600', None, 'Credit Bureau Team', 'MONTHLY'],
            
            # Geographic Features
            ['Province', 'PROVINCE', 'Tỉnh/thành phố của khách hàng', 'CATEGORICAL', 'NOMINAL', 'CORE_BANKING', 'GEOGRAPHIC', 'Vị trí địa lý có thể ảnh hưởng đến rủi ro tín dụng', False, False, 'UNKNOWN', None, None, '["HANOI", "HCMC", "DANANG", "HAIPHONG", "OTHERS"]', 'Customer Data Team', 'STATIC'],
            ['Urban vs Rural', 'URBAN_RURAL', 'Phân loại đô thị/nông thôn', 'CATEGORICAL', 'BINARY', 'CORE_BANKING', 'GEOGRAPHIC', 'Phân loại đô thị/nông thôn ảnh hưởng đến khả năng tiếp cận dịch vụ', False, False, 'UNKNOWN', None, None, '["URBAN", "RURAL"]', 'Customer Data Team', 'STATIC'],
            
            # Temporal Features
            ['Application Date', 'APP_DATE', 'Ngày đăng ký khoản vay', 'DATE', 'TEMPORAL', 'CORE_BANKING', 'TEMPORAL', 'Ngày đăng ký có thể ảnh hưởng đến điều kiện thị trường', False, False, None, None, None, None, 'Customer Data Team', 'STATIC'],
            ['Season', 'SEASON', 'Mùa trong năm', 'CATEGORICAL', 'CYCLICAL', 'DATA_WAREHOUSE', 'TEMPORAL', 'Mùa trong năm có thể ảnh hưởng đến hành vi tiêu dùng', False, False, None, None, None, '["SPRING", "SUMMER", "AUTUMN", "WINTER"]', 'Risk Analytics Team', 'MONTHLY'],
            
            # External Features
            ['GDP Growth Rate', 'GDP_GROWTH', 'Tỷ lệ tăng trưởng GDP', 'NUMERIC', 'CONTINUOUS', 'EXTERNAL_DATA', 'EXTERNAL', 'Tỷ lệ tăng trưởng GDP ảnh hưởng đến điều kiện kinh tế tổng thể', False, False, None, '-10', '20', None, 'Economic Research Team', 'QUARTERLY'],
            ['Unemployment Rate', 'UNEMPLOYMENT_RATE', 'Tỷ lệ thất nghiệp', 'NUMERIC', 'CONTINUOUS', 'EXTERNAL_DATA', 'EXTERNAL', 'Tỷ lệ thất nghiệp ảnh hưởng đến khả năng trả nợ', False, False, None, '0', '50', None, 'Economic Research Team', 'MONTHLY'],
            ['Interest Rate Environment', 'INT_RATE_ENV', 'Môi trường lãi suất', 'CATEGORICAL', 'ORDINAL', 'EXTERNAL_DATA', 'EXTERNAL', 'Môi trường lãi suất ảnh hưởng đến chi phí vốn', False, False, None, None, None, '["LOW", "MEDIUM", "HIGH", "VERY_HIGH"]', 'Economic Research Team', 'MONTHLY']
        ]
    }
}

def create_excel_template(table_name, config, output_dir='templates'):
    """Create Excel template with formatting and validation"""
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = table_name.replace('_', ' ').title()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    required_font = Font(bold=True, color="FF0000")
    info_font = Font(italic=True, color="808080")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws['A1'] = f"Template Upload - {table_name.replace('_', ' ').title()}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:Z1')
    
    # Add instructions
    ws['A3'] = "Hướng dẫn sử dụng:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "1. Các cột có màu đỏ là bắt buộc phải điền"
    ws['A5'] = "2. Các cột có dropdown chỉ được chọn từ danh sách có sẵn"
    ws['A6'] = "3. Định dạng ngày tháng: YYYY-MM-DD"
    ws['A7'] = "4. Giá trị boolean: 1 (True) hoặc 0 (False)"
    ws['A8'] = "5. Không xóa hoặc thay đổi tên cột"
    
    # Add column headers starting from row 10
    start_row = 10
    for col_idx, col_config in enumerate(config['columns'], 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = col_config['name']
        
        # Apply header styling
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Mark required columns in red
        if col_config.get('required', False):
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add description in row below
        desc_cell = ws.cell(row=start_row + 1, column=col_idx)
        desc_cell.value = col_config.get('description', '')
        desc_cell.font = info_font
        desc_cell.border = thin_border
    
    # Add sample data starting from row 12
    sample_start_row = start_row + 2
    if 'sample_data' in config:
        for row_idx, sample_row in enumerate(config['sample_data'], sample_start_row):
            for col_idx, value in enumerate(sample_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
    
    # Add data validation
    for col_idx, col_config in enumerate(config['columns'], 1):
        if col_config['type'] == 'dropdown':
            # Create dropdown validation
            dv = DataValidation(type="list", formula1=f'"{",".join(col_config["options"])}"', allow_blank=True)
            dv.add(f'{chr(64 + col_idx)}{sample_start_row}:{chr(64 + col_idx)}1000')
            ws.add_data_validation(dv)
        
        elif col_config['type'] == 'date':
            # Add date validation hint
            hint_cell = ws.cell(row=start_row + 1, column=col_idx)
            hint_cell.value = f"{col_config.get('description', '')} (YYYY-MM-DD)"
        
        elif col_config['type'] == 'boolean':
            # Create boolean dropdown
            dv = DataValidation(type="list", formula1='"1,0,True,False"', allow_blank=True)
            dv.add(f'{chr(64 + col_idx)}{sample_start_row}:{chr(64 + col_idx)}1000')
            ws.add_data_validation(dv)
    
    # Adjust column widths
    for col_idx, col_config in enumerate(config['columns'], 1):
        max_width = max(len(col_config['name']), len(col_config.get('description', '')), 15)
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_width + 2, 50)
    
    # Save template
    output_file = os.path.join(output_dir, f"{table_name}_template.xlsx")
    wb.save(output_file)
    print(f"Created template: {output_file}")
    
    return output_file

def create_sample_data_file(table_name, config, output_dir='sample_data'):
    """Create sample data file"""
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Create DataFrame from sample data
    if 'sample_data' in config:
        columns = [col['name'] for col in config['columns']]
        df = pd.DataFrame(config['sample_data'], columns=columns)
        
        # Save to Excel
        output_file = os.path.join(output_dir, f"{table_name}_sample.xlsx")
        df.to_excel(output_file, index=False, sheet_name='Sample Data')
        
        # Add formatting
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            # Create instructions sheet
            instructions_df = pd.DataFrame({
                'Field': ['Instructions', 'Required Fields', 'Data Types', 'Validation Rules'],
                'Description': [
                    'This file contains sample data for reference. Replace with your actual data.',
                    ', '.join([col['name'] for col in config['columns'] if col.get('required', False)]),
                    ', '.join([f"{col['name']}: {col['type']}" for col in config['columns']]),
                    'Follow the validation rules specified in the template file.'
                ]
            })
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
        
        print(f"Created sample data: {output_file}")
        return output_file
    
    return None

def main():
    """Main function to create all templates"""
    print("Creating Excel templates for Model Registry...")
    
    for table_name, config in TEMPLATE_CONFIGS.items():
        print(f"\nProcessing {table_name}...")
        
        # Create template
        template_file = create_excel_template(table_name, config)
        
        # Create sample data
        sample_file = create_sample_data_file(table_name, config)
        
        print(f"✓ {table_name}: Template and sample data created")
    
    print(f"\nAll templates created successfully!")
    print(f"Templates location: {os.path.abspath('templates')}")
    print(f"Sample data location: {os.path.abspath('sample_data')}")

if __name__ == "__main__":
    main() 